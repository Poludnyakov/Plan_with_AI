import io
from datetime import date
from unittest.mock import AsyncMock, patch

import pytest
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from database import Base
from schedule_document_service import (
    DocumentChunkExtraction,
    ParsedDocumentChunk,
    ScheduleDocumentClarification,
    _deterministic_scoped_slots,
    _ocr_text_from_payload,
    _scoped_table_weekdays,
    _table_selection_candidates,
    _verify_source_support,
    _verify_weekday_coverage,
    _clean_json,
    _validated_slots,
    document_chunks,
    extract_document_text,
    parse_schedule_document,
)
from schedule_models import ScheduleSeries
from schedule_service import confirm_import, create_import_draft, set_draft_range


@pytest.mark.anyio
async def test_xlsx_extraction_reads_every_sheet_and_nonempty_row():
    workbook = Workbook()
    first = workbook.active
    first.title = "Смены"
    first.append(["Сотрудник", "Дата", "Время"])
    first.append(["Анна", "03.09.2026", "09:00-18:00"])
    second = workbook.create_sheet("Тренировки")
    second.append(["Йога", "пятница", "19:00-20:00"])
    stream = io.BytesIO()
    workbook.save(stream)

    text = await extract_document_text(stream.getvalue(), "all-schedules.xlsx")

    assert "=== ЛИСТ: Смены ===" in text
    assert "Анна\t03.09.2026\t09:00-18:00" in text
    assert "=== ЛИСТ: Тренировки ===" in text
    assert "Йога\tпятница\t19:00-20:00" in text


@pytest.mark.anyio
async def test_xlsx_extraction_expands_merged_headers_for_llm_context():
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "Группа ПИ-241"
    sheet.append(["Понедельник", "09:00", "Математика"])
    stream = io.BytesIO()
    workbook.save(stream)

    text = await extract_document_text(stream.getvalue(), "merged.xlsx")

    assert "Группа ПИ-241\tГруппа ПИ-241\tГруппа ПИ-241" in text
    assert "Понедельник\t09:00\tМатематика" in text


def test_document_response_normalizes_wrappers_and_underscored_nested_keys():
    payload = _clean_json(
        '{"_embedded":{"_slots":[{"_title":"Смена","_weekday":null,'
        '"_occurrence_date":"2026-09-03","_start_time":"09:00",'
        '"_end_time":"18:00","_description":"офис"}]}}'
    )
    parsed = DocumentChunkExtraction.model_validate(payload)
    assert parsed.slots[0].occurrence_date == date(2026, 9, 3)
    assert parsed.slots[0].weekday == 3


def test_model_response_keeps_good_rows_when_one_row_is_broken():
    raw = '''Мой результат:
    {"slots":[
      {"title":"Смена","occurrence_date":"2026-09-03","start_time":"09:00","end_time":"18:00"},
      {"title":"Сломанная строка","weekday":2,"start_time":"18:00","end_time":"09:00"}
    ]}'''
    slots = _validated_slots(raw)
    assert [item["title"] for item in slots] == ["Смена"]


def test_large_table_chunks_repeat_column_context():
    text = "Сотрудник\tДата\tВремя\n" + "\n".join(
        f"Анна\t{index:02d}.09.2026\t09:00-18:00" for index in range(1, 21)
    )
    with patch("schedule_document_service.CHUNK_CHARS", 160), \
         patch("schedule_document_service.CHUNK_OVERLAP", 30):
        chunks = document_chunks(text)
    assert len(chunks) > 1
    assert all("Сотрудник\tДата\tВремя" in chunk for chunk in chunks[1:])


@pytest.mark.anyio
async def test_document_parser_processes_all_chunks_and_keeps_exact_and_recurring_rows():
    answers = [
        [{
            "title": "Смена", "weekday": 3, "occurrence_date": "2026-09-03",
            "start_time": "09:00", "end_time": "18:00", "description": "",
            "week_pattern": "every", "confidence": 0.96,
        }],
        [{
            "title": "Тренировка", "weekday": 4, "occurrence_date": None,
            "start_time": "19:00", "end_time": "20:00", "description": "",
            "week_pattern": "every", "confidence": 0.91,
        }],
    ]
    with patch("schedule_document_service.settings.API_KEY", "key"), \
         patch("schedule_document_service.settings.FOLDER_ID", "folder"), \
         patch("schedule_document_service.settings.YANDEX_DOCUMENT_READER_MODEL", "strong/latest"), \
         patch("schedule_document_service.settings.YANDEX_DOCUMENT_NORMALIZER_MODEL", "lite/latest"), \
         patch("schedule_document_service.extract_document_text", new_callable=AsyncMock, return_value="Четверг Смена 09:00 18:00\nПятница Тренировка 19:00 20:00"), \
         patch("schedule_document_service.document_chunks", return_value=["part one", "part two"]), \
         patch("schedule_document_service._parse_chunk", new_callable=AsyncMock, side_effect=answers) as parse, \
         patch("schedule_document_service._normalize_candidates", new_callable=AsyncMock, side_effect=lambda rows, *_: rows) as normalize:
        result = await parse_schedule_document(
            b"content", "schedule.xlsx", "выбери Анну с 1 сентября по 1 декабря"
        )

    assert parse.await_count == 2
    assert all(call.args[2] == "strong/latest" for call in parse.await_args_list)
    normalize.assert_not_awaited()
    assert result["chunks_processed"] == 2
    assert result["pipeline"] == "strong_reader_checked"
    assert {slot["title"] for slot in result["slots"]} == {"Смена", "Тренировка"}


@pytest.mark.anyio
async def test_image_uses_strong_multimodal_reader_before_ocr():
    rows = [{
        "title": "Тренировка", "weekday": 4, "occurrence_date": None,
        "start_time": "19:00", "end_time": "20:00", "description": "",
        "week_pattern": "every", "confidence": 0.95,
    }]
    with patch("schedule_document_service.settings.API_KEY", "key"), \
         patch("schedule_document_service.settings.FOLDER_ID", "folder"), \
         patch("schedule_document_service.settings.YANDEX_DOCUMENT_READER_MODEL", "strong/latest"), \
         patch("schedule_document_service.settings.YANDEX_DOCUMENT_NORMALIZER_MODEL", "lite/latest"), \
         patch("schedule_document_service._parse_image_direct", new_callable=AsyncMock, return_value=rows) as image_reader, \
         patch("schedule_document_service.extract_document_text", new_callable=AsyncMock, return_value="Пятница Тренировка 19:00 20:00") as extract_text, \
         patch("schedule_document_service._parse_chunk", new_callable=AsyncMock, return_value=rows), \
         patch("schedule_document_service._normalize_candidates", new_callable=AsyncMock, side_effect=lambda values, *_: values):
        result = await parse_schedule_document(
            b"image", "schedule.png", "тренировки с 1 сентября по 1 декабря"
        )

    image_reader.assert_awaited_once()
    assert image_reader.await_args.args[3] == "strong/latest"
    extract_text.assert_not_awaited()
    assert result["slots"][0]["title"] == "Тренировка"


@pytest.mark.anyio
async def test_empty_strong_reader_is_not_reinterpreted_by_lite_model():
    with patch("schedule_document_service.settings.API_KEY", "key"), \
         patch("schedule_document_service.settings.FOLDER_ID", "folder"), \
         patch("schedule_document_service.settings.YANDEX_DOCUMENT_READER_MODEL", "strong/latest"), \
         patch("schedule_document_service.settings.YANDEX_DOCUMENT_NORMALIZER_MODEL", "lite/latest"), \
         patch("schedule_document_service.extract_document_text", new_callable=AsyncMock, return_value="Вторник Дежурство 08:00 12:00"), \
         patch("schedule_document_service._parse_chunk", new_callable=AsyncMock, return_value=[]) as parse, \
         patch("schedule_document_service._normalize_candidates", new_callable=AsyncMock, side_effect=lambda values, *_: values):
        with pytest.raises(ValueError, match="подтверждаемые строки"):
            await parse_schedule_document(
                b"table", "schedule.csv", "дежурства с 1 сентября по 1 декабря"
            )

    assert parse.await_count == 1
    assert parse.await_args.args[2] == "strong/latest"


@pytest.mark.anyio
async def test_parser_without_prompt_accepts_exact_dates():
    rows = [{
        "title": "Математика", "weekday": 4, "occurrence_date": "2026-09-04",
        "start_time": "09:00", "end_time": "10:30", "description": "",
        "week_pattern": "every", "confidence": 0.96,
    }]
    with patch("schedule_document_service.settings.API_KEY", "key"), \
         patch("schedule_document_service.settings.FOLDER_ID", "folder"), \
         patch("schedule_document_service.extract_document_text", new_callable=AsyncMock, return_value="04.09.2026 Математика 09:00 10:30"), \
         patch("schedule_document_service._parse_chunk", new_callable=AsyncMock, return_value=rows) as parse:
        result = await parse_schedule_document(b"table", "schedule.csv", "")

    parse.assert_awaited_once()
    assert result["requires_range"] is False
    assert result["slots"][0]["occurrence_date"] == "2026-09-04"


def test_verifier_rejects_hallucinated_subject_but_keeps_source_title():
    source = "Понедельник 09:00 Математика\nСреда 10:00 Физика"
    rows = [
        {"title": "Математика", "weekday": 0},
        {"title": "Биология", "weekday": 2},
    ]

    accepted, rejected = _verify_source_support(rows, source)

    assert [item["title"] for item in accepted] == ["Математика"]
    assert rejected == ["Биология"]


def test_verifier_blocks_partial_week_instead_of_importing_two_days():
    source = (
        "Понедельник Математика\nВторник Физика\nСреда Химия\n"
        "Четверг История\nПятница Литература"
    )
    with pytest.raises(ValueError, match="только часть расписания"):
        _verify_weekday_coverage(
            [{"title": "Физика", "weekday": 1}, {"title": "Химия", "weekday": 2}],
            source,
        )


def test_multigroup_xls_coverage_uses_only_requested_group_column():
    source = "\n".join([
        "=== ЛИСТ: Расписание ===",
        "R1\tДень\tПИ-241\tПИ-242",
        "R2\tПонедельник\tМатематика\tБиология",
        "R3\tВторник\tФизика\tХимия",
        "R4\tСреда\tИстория\tЛитература",
        "R5\tЧетверг\t—\tИнформатика",
        "R6\tПятница\t—\tФилософия",
        "R7\tСуббота\t—\tПрактикум",
    ])

    days, labels, evidence = _scoped_table_weekdays(
        source, "группа ПИ-241 с 1 сентября по 1 декабря"
    )

    assert days == {0, 1, 2}
    assert labels == ["ПИ-241"]
    assert "Математика" in evidence
    assert "Биология" not in evidence
    assert _verify_weekday_coverage(
        [{"weekday": 0}, {"weekday": 1}, {"weekday": 2}],
        source,
        expected_days=days,
    )[0] == {0, 1, 2}

    rows, unresolved = _deterministic_scoped_slots(evidence, labels)
    assert rows == []
    assert unresolved == 3


@pytest.mark.anyio
async def test_unambiguous_group_projection_uses_strong_reader():
    source = "\n".join([
        "R1\tДень\tВремя\tПИ-241\tПИ-242",
        "R2\tПонедельник\t09:00-10:30\tМатематика\tБиология",
        "R3\tВторник\t10:00-11:30\tФизика\tХимия",
        "R4\tСреда\t11:00-12:30\tИстория\tЛитература",
        "R5\tЧетверг\t12:00-13:30\t—\tФилософия",
    ])
    rows = [
        {"title": "Математика", "weekday": 0, "start_time": "09:00", "end_time": "10:30"},
        {"title": "Физика", "weekday": 1, "start_time": "10:00", "end_time": "11:30"},
        {"title": "История", "weekday": 2, "start_time": "11:00", "end_time": "12:30"},
    ]
    with patch("schedule_document_service.settings.API_KEY", "key"), \
         patch("schedule_document_service.settings.FOLDER_ID", "folder"), \
         patch("schedule_document_service.extract_document_text", new_callable=AsyncMock, return_value=source), \
         patch("schedule_document_service._parse_chunk", new_callable=AsyncMock, return_value=rows) as parse:
        result = await parse_schedule_document(
            b"xls", "schedule.xls",
            "группа ПИ-241 с 1 сентября по 1 декабря",
        )

    parse.assert_awaited_once()
    assert result["pipeline"] == "strong_reader_checked"
    assert [item["title"] for item in result["slots"]] == [
        "Математика", "Физика", "История"
    ]

def test_multigroup_xls_rejects_missing_day_inside_requested_group():
    source = "\n".join([
        "R1\tДень\tПИ-241\tПИ-242",
        "R2\tПонедельник\tМатематика\tБиология",
        "R3\tВторник\tФизика\tХимия",
        "R4\tСреда\tИстория\tЛитература",
        "R5\tЧетверг\tИнформатика\tФилософия",
    ])
    days, _, _ = _scoped_table_weekdays(source, "расписание ПИ-241")

    with pytest.raises(ValueError, match="только часть расписания"):
        _verify_weekday_coverage(
            [{"weekday": 0}, {"weekday": 1}, {"weekday": 2}],
            source,
            expected_days=days,
        )


def test_ocr_preserves_line_coordinates_for_table_layout():
    texts = _ocr_text_from_payload({
        "textAnnotation": {
            "fullText": "Понедельник Математика 09:00",
            "blocks": [{"lines": [
                {"text": "Математика", "boundingBox": {"vertices": [{"x": 200, "y": 50}]}},
                {"text": "Понедельник", "boundingBox": {"vertices": [{"x": 10, "y": 50}]}},
            ]}],
        }
    })

    assert "Y00050 X00010\tПонедельник" in texts[0]
    assert texts[0].index("Понедельник") < texts[0].index("Математика")


@pytest.mark.anyio
async def test_exact_date_row_is_imported_only_on_that_date():
    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    sessions = async_sessionmaker(engine, expire_on_commit=False)
    async with engine.begin() as connection:
        await connection.run_sync(Base.metadata.create_all)
    async with sessions() as db:
        draft = await create_import_draft(db, "telegram", 9001, {
            "confidence": 0.96,
            "slots": [{
                "title": "Смена", "weekday": 3, "occurrence_date": "2026-09-03",
                "start_time": "09:00", "end_time": "18:00", "description": "офис",
                "week_pattern": "every", "confidence": 0.96,
            }],
        })
        assert draft.status == "ready"
        assert draft.valid_from == draft.valid_until == date(2026, 9, 3)
        assert await confirm_import(db, "telegram", 9001, draft.id) == ("created", 1)
        rows = (await db.execute(select(ScheduleSeries))).scalars().all()
        assert len(rows) == 1
        assert rows[0].valid_from == rows[0].valid_until == date(2026, 9, 3)
    await engine.dispose()



def test_header_candidates_do_not_treat_subject_codes_as_groups():
    source = "\n".join([
        "R1\tДень\tВремя\tПредмет",
        "R2\tПонедельник\t09:00-10:30\tБД-101",
        "R3\tВторник\t10:40-12:10\tОП-202",
    ])

    assert _table_selection_candidates(source) == []


@pytest.mark.anyio
async def test_multigroup_source_asks_after_strong_reader_detects_ambiguity():
    source = "\n".join([
        "=== ЛИСТ: Расписание ===",
        "R1\tДень\tВремя\tПИ-241\tПИ-242",
        "R2\tПонедельник\t09:00-10:30\tМатематика\tБиология",
        "R3\tВторник\t10:40-12:10\tФизика\tХимия",
    ])
    result = ParsedDocumentChunk(
        slots=[],
        clarification="Нужно выбрать одну группу: ПИ-241 или ПИ-242.",
        choices=["ПИ-241", "ПИ-242"],
    )
    with patch("schedule_document_service.settings.API_KEY", "key"), \
         patch("schedule_document_service.settings.FOLDER_ID", "folder"), \
         patch("schedule_document_service.extract_document_text", new_callable=AsyncMock, return_value=source), \
         patch("schedule_document_service._parse_chunk", new_callable=AsyncMock, return_value=result) as parse:
        with pytest.raises(ScheduleDocumentClarification, match="выбрать одну группу") as error:
            await parse_schedule_document(b"xls", "schedule.xls", "")

    parse.assert_awaited_once()
    assert error.value.choices == ["ПИ-241", "ПИ-242"]


@pytest.mark.anyio
async def test_reader_budget_stops_unknown_large_source_before_partial_import():
    source = "Понедельник Раздел 09:00 10:00"
    with patch("schedule_document_service.settings.API_KEY", "key"), \
         patch("schedule_document_service.settings.FOLDER_ID", "folder"), \
         patch("schedule_document_service.extract_document_text", new_callable=AsyncMock, return_value=source), \
         patch("schedule_document_service.document_chunks", return_value=["part"] * 5), \
         patch("schedule_document_service._parse_chunk", new_callable=AsyncMock) as parse:
        with pytest.raises(ScheduleDocumentClarification, match="слишком много"):
            await parse_schedule_document(
                b"large", "schedule.txt", "с 1 сентября по 1 декабря"
            )

    parse.assert_not_awaited()



@pytest.mark.anyio
async def test_magic_detects_xlsx_when_messenger_drops_filename_extension():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["День", "Время", "Предмет"])
    sheet.append(["Понедельник", "09:00-10:30", "Математика"])
    stream = io.BytesIO()
    workbook.save(stream)

    text = await extract_document_text(stream.getvalue(), "document")

    assert "Понедельник" in text
    assert "Математика" in text
